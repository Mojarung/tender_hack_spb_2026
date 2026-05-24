"""НМЦК (Начальная Максимальная Цена Контракта) calculator + Excel
export — Приложение №1 to 44-ФЗ, метод сопоставимых рыночных цен.

Statistics follow standard MED-РФ guidance:
  Средняя арифметическая  μ = Σx / n
  Среднеквадратическое отклонение  σ = √(Σ(x − μ)² / (n − 1))
  Коэффициент вариации  V = σ / μ × 100 %

V ≤ 33 %  → выборка однородна, НМЦК = μ
V > 33 %  → выборка неоднородна, нужно расширить (4+ КП) или применить
            метод тарифа / проектно-сметный.

The workbook is built for a non-specialist reader (a buyer at a school
or a regional ministry who runs a tender once a quarter): the first
page is a single glanceable summary card with the final НМЦК, the
verdict, and an explanation of every number. The five КП used in the
formula sit below. A second sheet lists every offer we found so the
analyst can verify that the cheapest five are reasonable.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pricepulse.domain.enums import SourceKind
from pricepulse.domain.models import ProductOffer

# Friendly Russian labels — what the buyer expects to see, not raw enum values.
_SOURCE_LABEL: dict[str, str] = {
    "wb": "Wildberries",
    "ozon": "Ozon",
    "ya_market": "Яндекс.Маркет",
    "runet": "Магазины Рунета",
}

# Cap on КП used in the formula. 44-ФЗ requires ≥3; using all 30+ offers
# pulls the average toward outliers (e.g. one absurdly cheap counterfeit)
# and inflates the variation coefficient. Standard госзакупки practice
# is "3-5 reliable КП"; we lean on 5.
_KP_FORMULA_LIMIT = 5

# ─────────────────────────────────────────── pure-Python core ─────────


@dataclass(frozen=True)
class NmckStats:
    n_offers: int
    mean: Decimal
    stdev: Decimal
    cv_pct: float
    homogeneous: bool
    nmck_per_unit: Decimal
    sources_used: list[str]


def compute(offers: list[ProductOffer]) -> NmckStats | None:
    """Pick the cheapest offer from each unique seller (or source), trim
    to at most 5 КП, compute statistics. Returns None when fewer than
    three valid КП — 44-ФЗ requires three minimum."""
    chosen = _select_kp(offers)
    if len(chosen) < 3:
        return None
    prices = [float(o.price) for o in chosen]
    mu = sum(prices) / len(prices)
    var = sum((p - mu) ** 2 for p in prices) / (len(prices) - 1)
    sigma = math.sqrt(var)
    v = (sigma / mu * 100) if mu > 0 else 0.0
    return NmckStats(
        n_offers=len(chosen),
        mean=Decimal(str(round(mu, 2))),
        stdev=Decimal(str(round(sigma, 2))),
        cv_pct=round(v, 2),
        homogeneous=v <= 33.0,
        nmck_per_unit=Decimal(str(round(mu, 2))),
        sources_used=[_seller_label(o) for o in chosen],
    )


def _select_kp(offers: list[ProductOffer]) -> list[ProductOffer]:
    """Pick the КП that a procurement officer would actually defend
    in tender paperwork — not just "cheapest from each seller", because
    that gives the formula whatever counterfeit listing is currently
    sitting at the top of WB.

    Strategy:
      1. Drop offers without a price.
      2. Dedup by seller, keep the cheapest from each (a seller's own
         listings rarely diverge meaningfully).
      3. Compute a rough market median across the dedup'd pool. Drop
         anything below median × 0.5 (almost certainly a counterfeit
         or wrong-SKU) or above median × 2 (premium bundle / wrong
         category match).
      4. Score the survivors by "trustworthiness":
            trust = log10(reviews + 1) × rating
         Offers with no rating get 0.5×rating mid-tier credit so they
         don't get hard-zeroed (Yandex SERP path often lacks rating).
      5. Take the top _KP_FORMULA_LIMIT by trust score, but break ties
         and fill gaps with proximity-to-median so we still end up with
         realistic prices, not just popular ones.
      6. Fallback — if filtering leaves <3 КП (a rare query with bad
         coverage), relax and fall back to the cheapest-per-seller
         pool so the report still renders."""
    by_seller: dict[str, ProductOffer] = {}
    for o in offers:
        if o.price is None or o.price <= 0:
            continue
        key = (o.seller or _source_value(o)).strip().lower()
        cur = by_seller.get(key)
        if cur is None or o.price < cur.price:
            by_seller[key] = o
    pool = list(by_seller.values())
    if len(pool) < 3:
        return sorted(pool, key=lambda o: o.price)

    prices = sorted(float(o.price) for o in pool)
    mid = len(prices) // 2
    market_median = prices[mid] if len(prices) % 2 else (prices[mid - 1] + prices[mid]) / 2

    def in_band(o: ProductOffer) -> bool:
        p = float(o.price)
        return market_median * 0.5 <= p <= market_median * 2.0

    realistic = [o for o in pool if in_band(o)]
    if len(realistic) < 3:
        # Filter was too aggressive — fall back to the full dedup pool
        # sorted by closeness to median so the formula still gets
        # representative prices.
        realistic = sorted(pool, key=lambda o: abs(float(o.price) - market_median))

    def trust(o: ProductOffer) -> float:
        reviews = float(o.reviews_count or 0)
        # log10(reviews+1) keeps it well-behaved at 0; rating defaults to
        # 4.0 when missing so absent metadata doesn't penalise sellers
        # that simply didn't expose rating in the scrape.
        rating = float(o.rating) if o.rating is not None else 4.0
        return math.log10(reviews + 1.0) * rating

    realistic.sort(key=lambda o: (-trust(o), abs(float(o.price) - market_median)))
    return realistic[:_KP_FORMULA_LIMIT]


def _source_value(o: ProductOffer) -> str:
    src = o.source
    if isinstance(src, SourceKind):
        return src.value
    return str(src)


def _source_label(o: ProductOffer) -> str:
    return _SOURCE_LABEL.get(_source_value(o), _source_value(o))


def _seller_label(o: ProductOffer) -> str:
    """Best-effort "shop name" — seller if present, else marketplace name."""
    if o.seller and o.seller.strip():
        return o.seller.strip()
    return _source_label(o)


# ─────────────────────────────────────────── Excel renderer ──────────


_THIN = Side(border_style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_VERDICT_OK = PatternFill("solid", fgColor="C6EFCE")
_VERDICT_BAD = PatternFill("solid", fgColor="FFC7CE")
_SUMMARY_BG = PatternFill("solid", fgColor="F2F7FC")
_NMCK_BG = PatternFill("solid", fgColor="FFF2CC")
_NMCK_FONT = Font(bold=True, size=18, color="1F4E79")
_LABEL_FONT = Font(bold=True, color="404040")
_CAPTION_FONT = Font(italic=True, color="606060", size=9)


def _put(ws, row: int, col: int, value, *, font=None, fill=None,
         border=False, align=None, num_format=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if border:
        cell.border = _BORDER
    if align is not None:
        cell.alignment = align
    if num_format is not None:
        cell.number_format = num_format


def _render_summary_sheet(
    ws, *, query: str, offers: list[ProductOffer],
    stats: NmckStats, quantity: int, total_found: int,
) -> None:
    """Page 1 — single-glance dashboard for the buyer."""
    ws.title = "НМЦК"
    widths = [4, 32, 18, 14, 18, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    centred = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    today = datetime.now(tz=UTC).strftime("%d.%m.%Y")

    # ─── Title ───
    ws.merge_cells("A1:F1")
    _put(ws, 1, 1, "Сколько закладывать на закупку",
         font=Font(bold=True, size=16, color="1F4E79"), align=centred)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    _put(ws, 2, 1,
         "Готовое обоснование цены контракта по 44-ФЗ — можно сразу "
         "вставлять в закупочную документацию",
         font=_CAPTION_FONT, align=centred)

    # ─── What this document is — short, plain Russian ───
    ws.merge_cells("A4:F6")
    _put(ws, 4, 1,
         "Когда государственное учреждение что-то покупает, оно сначала "
         "должно посчитать «справедливую» цену по рынку — это и есть НМЦК "
         "(начальная максимальная цена контракта). Мы взяли цены этого "
         "товара в нескольких крупных интернет-магазинах и посчитали "
         "среднюю. Эту цифру можно использовать как стартовую на торгах.",
         font=Font(size=11), align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[4].height = 60

    # ─── Big summary card ───
    summary_row = 8
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    _put(ws, summary_row, 1, "Рекомендуемая цена за 1 шт.",
         font=Font(bold=True, size=11, color="404040"),
         fill=_NMCK_BG, align=centred, border=True)
    ws.merge_cells(f"D{summary_row}:F{summary_row}")
    _put(ws, summary_row, 4, float(stats.mean), font=_NMCK_FONT,
         fill=_NMCK_BG, align=centred, border=True,
         num_format='#,##0.00 "₽"')
    ws.row_dimensions[summary_row].height = 42

    if quantity > 1:
        total_row = summary_row + 1
        ws.merge_cells(f"A{total_row}:C{total_row}")
        _put(ws, total_row, 1, f"Итого за партию ({quantity} шт.)",
             font=Font(bold=True, size=11, color="404040"),
             fill=_NMCK_BG, align=centred, border=True)
        ws.merge_cells(f"D{total_row}:F{total_row}")
        _put(ws, total_row, 4, float(stats.mean) * quantity,
             font=Font(bold=True, size=14, color="1F4E79"),
             fill=_NMCK_BG, align=centred, border=True,
             num_format='#,##0.00 "₽"')
        ws.row_dimensions[total_row].height = 26

    # ─── Verdict block ───
    verdict_row = summary_row + (3 if quantity > 1 else 2)
    fill = _VERDICT_OK if stats.homogeneous else _VERDICT_BAD
    verdict_text = (
        "✓ Цены в разных магазинах близки друг к другу — расчёту можно доверять"
        if stats.homogeneous
        else "⚠ Цены сильно разные (отличия больше трети) — лучше перепроверить вручную"
    )
    ws.merge_cells(f"A{verdict_row}:F{verdict_row}")
    _put(ws, verdict_row, 1, verdict_text,
         font=Font(bold=True, size=11), fill=fill, align=centred, border=True)
    ws.row_dimensions[verdict_row].height = 26

    # ─── Quick stats row — plain-language labels ───
    qs_row = verdict_row + 2
    quick = [
        ("Сколько магазинов проверили", str(stats.n_offers),
         "по закону нужно минимум 3"),
        ("Средняя цена", f"{stats.mean:,.0f} ₽".replace(",", " "),
         "её и берём как НМЦК"),
        ("Самый дешёвый магазин", f"{min(float(o.price) for o in _select_kp(offers)):,.0f} ₽".replace(",", " "),
         "ориентир «снизу»"),
        ("Разброс между ценами", f"{stats.cv_pct:.0f}%",
         "норма — не больше 33%"),
    ]
    # 2×2 grid of stat tiles: each tile = label / value / hint stacked
    # in 3 rows × 3 columns (so two tiles fit across A:F).
    for i, (label, value, hint) in enumerate(quick):
        # Place tiles in columns A-B, C-D, E-F by index 0,1; row stack of 3.
        # 4 tiles don't fit cleanly in 6 cols → do 2x2 grid instead.
        col0 = (i % 2) * 3 + 1
        row0 = qs_row + (i // 2) * 4
        ws.merge_cells(start_row=row0, start_column=col0,
                       end_row=row0, end_column=col0 + 2)
        _put(ws, row0, col0, label, font=_LABEL_FONT,
             fill=_SUMMARY_BG, align=Alignment(horizontal="left", indent=1),
             border=True)
        ws.merge_cells(start_row=row0 + 1, start_column=col0,
                       end_row=row0 + 1, end_column=col0 + 2)
        _put(ws, row0 + 1, col0, value,
             font=Font(bold=True, size=14),
             fill=_SUMMARY_BG, align=Alignment(horizontal="left", indent=1),
             border=True)
        ws.merge_cells(start_row=row0 + 2, start_column=col0,
                       end_row=row0 + 2, end_column=col0 + 2)
        _put(ws, row0 + 2, col0, hint,
             font=_CAPTION_FONT, fill=_SUMMARY_BG,
             align=Alignment(horizontal="left", indent=1), border=True)
        ws.row_dimensions[row0 + 2].height = 18

    # ─── Меta block ───
    meta_row = qs_row + 9
    _put(ws, meta_row, 1, "Что покупаем:", font=_LABEL_FONT)
    ws.merge_cells(start_row=meta_row, start_column=2,
                   end_row=meta_row, end_column=6)
    _put(ws, meta_row, 2, query, align=left_wrap)

    _put(ws, meta_row + 1, 1, "Дата проверки цен:", font=_LABEL_FONT)
    _put(ws, meta_row + 1, 2, today)
    _put(ws, meta_row + 2, 1, "Где искали:", font=_LABEL_FONT)
    ws.merge_cells(start_row=meta_row + 2, start_column=2,
                   end_row=meta_row + 2, end_column=6)
    _put(ws, meta_row + 2, 2,
         "Wildberries, Ozon, интернет-магазины Рунета (DNS, М.Видео, "
         "Эльдорадо и другие — через поиск Яндекса)",
         align=left_wrap)
    _put(ws, meta_row + 3, 1, "Найдено всего:", font=_LABEL_FONT)
    _put(ws, meta_row + 3, 2,
         f"{total_found} предложений · в расчёт взяли {stats.n_offers} "
         "самых надёжных (по рейтингу и количеству отзывов)")

    # ─── КП table — 5 rows used in the formula ───
    table_title_row = meta_row + 5
    ws.merge_cells(start_row=table_title_row, start_column=1,
                   end_row=table_title_row, end_column=6)
    _put(ws, table_title_row, 1,
         f"ЦЕНЫ, ВЗЯТЫЕ ДЛЯ РАСЧЁТА ({stats.n_offers} шт. — "
         "по одной из каждого магазина с лучшим рейтингом)",
         font=_LABEL_FONT, align=Alignment(horizontal="left"))

    hdr_row = table_title_row + 1
    headers = ["№", "Магазин", "Цена за 1 шт.", "Кол-во",
               "Итого", "Что именно покупаем"]
    for col, h in enumerate(headers, start=1):
        _put(ws, hdr_row, col, h, font=_HEADER_FONT, fill=_HEADER_FILL,
             border=True, align=centred)
    ws.row_dimensions[hdr_row].height = 30

    chosen = _select_kp(offers)
    first_data = hdr_row + 1
    for i, o in enumerate(chosen, start=1):
        r = first_data + i - 1
        _put(ws, r, 1, i, border=True, align=centred)
        _put(ws, r, 2, _seller_label(o), border=True, align=left_wrap)
        _put(ws, r, 3, float(o.price), border=True, align=right,
             num_format='#,##0.00 "₽"')
        _put(ws, r, 4, quantity, border=True, align=right)
        _put(ws, r, 5, f"=C{r}*D{r}", border=True, align=right,
             num_format='#,##0.00 "₽"')
        _put(ws, r, 6, o.name, border=True, align=left_wrap)
        ws.row_dimensions[r].height = 32

    last_data = first_data + len(chosen) - 1

    # ─── How the price was calculated — visible math, no jargon ───
    stats_row = last_data + 2
    ws.merge_cells(start_row=stats_row, start_column=1,
                   end_row=stats_row, end_column=6)
    _put(ws, stats_row, 1, "КАК ПОЛУЧИЛАСЬ ЦЕНА",
         font=_LABEL_FONT, align=Alignment(horizontal="left"))

    _put(ws, stats_row + 1, 1, "Средняя из цен выше",
         font=Font(bold=True, size=11),
         align=Alignment(horizontal="left", indent=1))
    ws.merge_cells(start_row=stats_row + 1, start_column=2,
                   end_row=stats_row + 1, end_column=4)
    _put(ws, stats_row + 1, 2, "← это и есть рекомендуемая цена",
         font=_CAPTION_FONT, align=Alignment(horizontal="left", indent=1))
    _put(ws, stats_row + 1, 5, f"=AVERAGE(C{first_data}:C{last_data})",
         font=Font(bold=True, size=11), border=True, align=right,
         num_format='#,##0.00 "₽"', fill=_NMCK_BG)

    _put(ws, stats_row + 2, 1, "Типичное отклонение от средней",
         font=_LABEL_FONT, align=Alignment(horizontal="left", indent=1))
    ws.merge_cells(start_row=stats_row + 2, start_column=2,
                   end_row=stats_row + 2, end_column=4)
    _put(ws, stats_row + 2, 2,
         "насколько обычно цена «гуляет» от средней",
         font=_CAPTION_FONT, align=Alignment(horizontal="left", indent=1))
    _put(ws, stats_row + 2, 5, f"=STDEV(C{first_data}:C{last_data})",
         border=True, align=right, num_format='#,##0.00 "₽"')

    _put(ws, stats_row + 3, 1, "Разброс между ценами",
         font=_LABEL_FONT, align=Alignment(horizontal="left", indent=1))
    ws.merge_cells(start_row=stats_row + 3, start_column=2,
                   end_row=stats_row + 3, end_column=4)
    _put(ws, stats_row + 3, 2,
         "если больше 33% — лучше проверить цены вручную",
         font=_CAPTION_FONT, align=Alignment(horizontal="left", indent=1))
    _put(ws, stats_row + 3, 5, f"=E{stats_row + 2}/E{stats_row + 1}*100",
         border=True, align=right, num_format='0.0"%"')

    # ─── Footer ───
    foot_row = stats_row + 5
    ws.merge_cells(start_row=foot_row, start_column=1,
                   end_row=foot_row + 1, end_column=6)
    _put(ws, foot_row, 1,
         f"Документ собран автоматически системой PricePulse · {today}. "
         "Все цены и ссылки на товары — на вкладке «Все источники». "
         "Любую цифру можно поменять прямо в этой таблице, итог пересчитается.",
         font=_CAPTION_FONT, align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[foot_row].height = 24

    ws.sheet_view.showGridLines = False


def _render_all_sources_sheet(ws, offers: list[ProductOffer]) -> None:
    """Page 2 — every offer we found, with friendly column names. The
    analyst can use this to sanity-check that the КП on page 1 are
    realistic (e.g. spot a counterfeit listing that dragged the
    average down)."""
    ws.title = "Все цены"
    widths = [4, 18, 28, 50, 16, 12, 14, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    centred = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:H1")
    _put(ws, 1, 1,
         "Все цены, которые мы нашли",
         font=Font(bold=True, size=14, color="1F4E79"),
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    _put(ws, 2, 1,
         "Здесь все предложения, которые удалось найти. На вкладке «НМЦК» "
         "мы взяли из них 5 самых надёжных — у магазинов с хорошим рейтингом "
         "и большим числом отзывов. Любую цену можно открыть по ссылке справа.",
         font=_CAPTION_FONT, align=Alignment(wrap_text=True, vertical="top"))
    ws.row_dimensions[2].height = 32

    headers = ["№", "Где продают", "Название магазина", "Что продают",
               "Цена", "Оценка", "Отзывов", "Перейти к товару"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        _put(ws, header_row, col, h, font=_HEADER_FONT, fill=_HEADER_FILL,
             border=True, align=centred)
    ws.row_dimensions[header_row].height = 26

    sortable = [o for o in offers if o.price is not None and o.price > 0]
    sortable.sort(key=lambda o: (_source_label(o), o.price))

    for i, o in enumerate(sortable, start=1):
        r = header_row + i
        _put(ws, r, 1, i, border=True, align=centred)
        _put(ws, r, 2, _source_label(o), border=True, align=left_wrap)
        _put(ws, r, 3, o.seller or "—", border=True, align=left_wrap)
        _put(ws, r, 4, o.name, border=True, align=left_wrap)
        _put(ws, r, 5, float(o.price), border=True, align=right,
             num_format='#,##0.00 "₽"')
        _put(ws, r, 6, float(o.rating) if o.rating is not None else "—",
             border=True, align=centred,
             num_format="0.0" if o.rating is not None else None)
        _put(ws, r, 7, o.reviews_count if o.reviews_count is not None else "—",
             border=True, align=centred)
        url_value = str(o.url) if o.url else ""
        if url_value:
            cell = ws.cell(row=r, column=8, value=url_value)
            cell.hyperlink = url_value
            cell.font = Font(color="1F4E79", underline="single", size=9)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            _put(ws, r, 8, "", border=True, align=left_wrap)
        ws.row_dimensions[r].height = 28

    # Auto-filter so the buyer can drill down per marketplace / seller.
    if sortable:
        ws.auto_filter.ref = f"A{header_row}:H{header_row + len(sortable)}"
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False


def to_excel(
    *,
    query: str,
    offers: list[ProductOffer],
    stats: NmckStats,
    quantity: int = 1,
) -> bytes:
    """Render the buyer-friendly two-page report. Returns xlsx bytes."""
    wb = Workbook()
    total_found = sum(1 for o in offers if o.price is not None and o.price > 0)
    _render_summary_sheet(
        wb.active, query=query, offers=offers, stats=stats,
        quantity=quantity, total_found=total_found,
    )
    second = wb.create_sheet()
    _render_all_sources_sheet(second, offers)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["NmckStats", "compute", "to_excel"]
